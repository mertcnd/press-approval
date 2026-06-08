import io, json, base64, os
from flask import Flask, request, send_file, Response
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Templates are embedded as base64 at deploy time — see build_app.py
PRESS_TEMPLATE_B64 = open(os.path.join(os.path.dirname(__file__), "press_template_b64.txt")).read().strip()
MAKEREADY_TEMPLATE_B64 = open(os.path.join(os.path.dirname(__file__), "makeready_template_b64.txt")).read().strip()

DOT_ROWS = [0,3,"5 lam",10,20,25,30,40,45,"50 lam",60,70,75,80,"lamination",100]
PULL_ROWS = [(9,10,13,14,15),(19,20,23,24,25),(29,30,33,34,35),(39,40,43,44,45),(49,50,53,54,55),(58,59,62,63,64)]

# ── PRESS APPROVAL Excel ──────────────────────────────────────────────────────

def fill_press_approval(data):
    wb = load_workbook(io.BytesIO(base64.b64decode(PRESS_TEMPLATE_B64)))
    ws = wb["Overview"]
    ws["B8"]=data.get("date",""); ws["B9"]=data.get("time","")
    ws["B11"]=data.get("printer",""); ws["B13"]=data.get("printerAddress","")
    ws["G11"]=data.get("printerAttendee",""); ws["G12"]=data.get("sgkAttendee",""); ws["G13"]=data.get("clientAttendee","")
    ws["B20"]=data.get("brand",""); ws["B21"]=data.get("variant","")
    ws["G20"]=data.get("clientProductRef",""); ws["G21"]=data.get("sgkJobNumber","")
    ws["B26"]=data.get("notes",""); ws["B36"]=data.get("jobPreview","")

    ws2 = wb["On Press"]
    for rows in PULL_ROWS:
        ws2.cell(row=rows[0],column=2).value=None; ws2.cell(row=rows[1],column=2).value=None
        for j in range(3): ws2.cell(row=rows[2+j],column=2).value=None
    for i,pull in enumerate(data.get("pulls",[])[:6]):
        rows=PULL_ROWS[i]
        ws2.cell(row=rows[0],column=2).value=pull.get("time","")
        ws2.cell(row=rows[1],column=2).value=pull.get("comments","")
        adjs=pull.get("adjustments",["","",""])
        for j in range(3): ws2.cell(row=rows[2+j],column=2).value=adjs[j] if j<len(adjs) else ""
    ap=data.get("approvedPull")
    ws2["B68"]=(ap+1) if ap is not None else ""; ws2["B69"]=data.get("approvalTime",""); ws2["B70"]=data.get("approvalComments","")

    ws3=wb["Print Sample Analysis"]
    ws3["B8"]=data.get("printSampleComments",""); ws3["B20"]=data.get("issuesAnalysis",""); ws3["B32"]=data.get("improvements","")
    colors=data.get("colors",[]); dotgain=data.get("dotgain",{}); density=data.get("density",{}); lab=data.get("lab",{})
    for r in range(43,64):
        for c in range(3,9): ws3.cell(row=r,column=c).value=None
    for ci,color in enumerate(colors):
        col=3+ci; cell=ws3.cell(row=43,column=col); cell.value=color["label"]
        hex_col=color["hex"].lstrip("#")
        if len(hex_col)==6:
            try: cell.fill=PatternFill("solid",fgColor="FF"+hex_col.upper())
            except: pass
        cell.font=Font(bold=True,size=12); cell.alignment=Alignment(horizontal="center",vertical="center")
    for ri,linear in enumerate(DOT_ROWS):
        for ci,color in enumerate(colors):
            vals=dotgain.get(color["id"],[])
            ws3.cell(row=44+ri,column=3+ci).value=vals[ri] if ri<len(vals) and vals[ri]!="" else None
    for ci,color in enumerate(colors):
        ws3.cell(row=60,column=3+ci).value=density.get(color["id"],"") or None
        lab_c=lab.get(color["id"],{})
        ws3.cell(row=61,column=3+ci).value=lab_c.get("L","") or None
        ws3.cell(row=62,column=3+ci).value=lab_c.get("a","") or None
        ws3.cell(row=63,column=3+ci).value=lab_c.get("b","") or None

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── MAKEREADY Excel ───────────────────────────────────────────────────────────

def fill_makeready(data):
    wb = load_workbook(io.BytesIO(base64.b64decode(MAKEREADY_TEMPLATE_B64)))
    ws = wb["Overview"]
    job=data.get("job",{}); history=data.get("history",{}); colors_d=data.get("colors",{})
    supplier=data.get("supplier",{}); client=data.get("client",{}); result_d=data.get("result",{})

    def w(coord, val):
        ws[coord] = val if val is not None else ""

    w("B8",  job.get("date",""))
    w("B9",  data.get("time","") or job.get("time",""))
    w("B11", job.get("supplier",""))
    w("B12", job.get("technique",""))
    w("B13", job.get("substrate",""))
    w("B17", job.get("brand",""))
    w("B18", job.get("variant",""))
    w("G18", job.get("jobNo",""))

    hp=history.get("hasPrevious")
    w("B24", "✓" if hp is True else "")
    w("F24", "✓" if hp is False else "")
    rsa=history.get("refSampleAvailable")
    w("B27", "✓" if rsa is True else "")
    w("F27", "✓" if rsa is False else "")
    w("C28", history.get("refSampleNo",""))
    w("C29", history.get("refSampleDate",""))
    mc=history.get("majorChange")
    w("B31", "✓" if mc is True else "")
    w("F31", "✓" if mc is False else "")
    w("C32", history.get("majorChangeDesc","") if mc is True else "")
    issues=history.get("prevIssues",[])
    if issues:
        lines = []
        for i,iss in enumerate(issues):
            n = iss.get("notified")
            iletildi = "Evet" if n is True else "Hayır" if n is False else "—"
            lines.append(str(i+1) + ". " + iss.get("note","") + " — İletildi: " + iletildi)
        w("B35", "\n".join(lines))

    gp=colors_d.get("gmgProfile")
    w("B43", "✓" if gp is True else "")
    w("F43", "✓" if gp is False else "")
    w("C44", colors_d.get("gmgProfileName","") if gp is True else "")
    spots=colors_d.get("spotColors",[]); spot_row=51
    for sc in spots[:6]:
        ws.cell(row=spot_row,column=1).value=sc.get("name","")
        ws.cell(row=spot_row,column=2).value=sc.get("pantone","")
        ws.cell(row=spot_row,column=3).value=sc.get("lab","")
        ws.cell(row=spot_row,column=4).value=sc.get("tolerance","")
        spot_row+=1

    sv=supplier.get("substrateConfirmed")
    w("B61","✓" if sv is True else ""); w("C61","✓" if sv is False else "")
    ws["E61"]=supplier.get("substrateNote") or ""
    tm=supplier.get("techniqueMatch")
    w("B64","✓" if tm is True else ""); w("C64","✓" if tm is False else "")
    ws["E64"]=supplier.get("techniqueNote") or ""
    ps=supplier.get("proofSent")
    w("B68","✓" if ps is True else ""); w("C68","✓" if ps is False else "")
    ws["E68"]=supplier.get("proofSentNote") or ""
    w("B72", supplier.get("proofStatus","") if ps is True else "")

    pa=client.get("proofApproved")
    w("B77","✓" if pa is True else ""); w("C77","✓" if pa is False else "")
    ws["E77"]=client.get("proofApprovedNote") or ""
    pin=client.get("prevIssuesNotified")
    w("B80","✓" if pin is True else ""); w("C80","✓" if pin is False else "")
    ws["E80"]=client.get("prevIssuesNotifiedNote") or ""
    w("C86", result_d.get("notes",""))

    # KARAR — seçilmeyeni görünmez yap (font rengi = arka plan rengi)
    decision=result_d.get("decision")
    rc=ws.cell(row=95,column=1)   # A95 — BASKIYA HAZIR (bg: yeşil FF92D050)
    nc=ws.cell(row=95,column=5)   # E95 — HAZIR DEĞİL   (bg: kırmızı FFFF0000)
    if decision=="ready":
        rc.font=Font(bold=True,size=20,color="FF1A6B1A")   # koyu yeşil yazı → görünür
        nc.font=Font(bold=True,size=20,color="FFFF0000")   # kırmızı yazı = kırmızı bg → görünmez
    elif decision=="notready":
        rc.font=Font(bold=True,size=20,color="FF92D050")   # yeşil yazı = yeşil bg → görünmez
        nc.font=Font(bold=True,size=20,color="FFFFFFFF")   # beyaz yazı → görünür

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── ANA SAYFA ─────────────────────────────────────────────────────────────────

HOME_HTML = """<!DOCTYPE html>
<html lang="tr">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SGK · Print Production</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#0f0f12;color:#e8e8f0;font-family:"DM Mono","Courier New",monospace;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}
.wrap{max-width:520px;width:100%}
.logo{font-size:11px;color:#9090a8;letter-spacing:4px;margin-bottom:8px}
h1{font-size:26px;font-weight:700;color:#e8c84a;margin-bottom:6px}
.sub{font-size:13px;color:#9090a8;margin-bottom:48px}
.cards{display:flex;flex-direction:column;gap:14px}
.card{display:block;background:#1c1c24;border:1px solid #2a2a36;border-radius:12px;padding:24px 28px;text-decoration:none;color:inherit;transition:border-color .2s,background .2s}
.card:hover{border-color:#e8c84a55;background:#1e1e28}
.card-label{font-size:10px;color:#e8c84a;letter-spacing:3px;font-weight:700;margin-bottom:8px}
.card-title{font-size:18px;font-weight:700;color:#e8e8f0;margin-bottom:6px}
.card-desc{font-size:12px;color:#9090a8;line-height:1.6}
.arrow{float:right;color:#e8c84a;font-size:20px;margin-top:-2px}
</style></head>
<body>
<div class="wrap">
  <div class="logo">SGK · PRINT PRODUCTION</div>
  <h1>Form Merkezi</h1>
  <p class="sub">Baskı süreçleri için dijital formlar</p>
  <div class="cards">
    <a href="/press-approval" class="card">
      <span class="arrow">→</span>
      <div class="card-label">BASKI ESNASI</div>
      <div class="card-title">Press Approval Report</div>
      <div class="card-desc">Pull kayıtları, renk ölçümleri, dot gain tablosu ve baskı onayı</div>
    </a>
    <a href="/makeready" class="card">
      <span class="arrow">→</span>
      <div class="card-label">BASKI ÖNCESİ</div>
      <div class="card-title">Makeready Checklist</div>
      <div class="card-desc">İş geçmişi, renk referansları, tedarikçi ve müşteri hazırlık kontrolü</div>
    </a>
  </div>
</div>
</body></html>"""

# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(HOME_HTML, mimetype="text/html")

@app.route("/press-approval")
def press_approval():
    path=os.path.join(os.path.dirname(__file__),"press_approval_report.html")
    return Response(open(path,"rb").read(), mimetype="text/html")

@app.route("/makeready")
def makeready_page():
    path=os.path.join(os.path.dirname(__file__),"makeready_checklist.html")
    return Response(open(path,"rb").read(), mimetype="text/html")

def cors(resp):
    resp.headers["Access-Control-Allow-Origin"]="*"
    resp.headers["Access-Control-Allow-Methods"]="POST,OPTIONS"
    resp.headers["Access-Control-Allow-Headers"]="Content-Type"
    return resp

@app.route("/generate", methods=["POST","OPTIONS"])
def generate():
    if request.method=="OPTIONS": return cors(Response())
    data=request.get_json()
    xlsx=fill_press_approval(data)
    brand=data.get("brand","Report").replace(" ","_")
    date=data.get("date","").replace(".","-")
    return cors(send_file(io.BytesIO(xlsx),as_attachment=True,
        download_name="Press_Approval_"+brand+"_"+date+".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

@app.route("/generate-makeready", methods=["POST","OPTIONS"])
def generate_makeready():
    if request.method=="OPTIONS": return cors(Response())
    data=request.get_json()
    xlsx=fill_makeready(data)
    job=data.get("job",{}); brand=job.get("brand","Report").replace(" ","_")
    date=data.get("date","").replace(".","-")
    return cors(send_file(io.BytesIO(xlsx),as_attachment=True,
        download_name="Makeready_"+brand+"_"+date+".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

# ── PDF GENERATİON ───────────────────────────────────────────────────────────

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

def generate_makeready_pdf(data):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=1.5*cm)
    DARK=rl_colors.HexColor("#0f0f12"); CARD=rl_colors.HexColor("#1c1c24"); ACCENT=rl_colors.HexColor("#e8c84a")
    TEXT=rl_colors.HexColor("#e8e8f0"); MUTED=rl_colors.HexColor("#9898b0"); OK_C=rl_colors.HexColor("#3ecf8e")
    OK_BG=rl_colors.HexColor("#1a3d2a"); NOK_C=rl_colors.HexColor("#e84a4a"); NOK_BG=rl_colors.HexColor("#3d1a1a")
    SEC_BG=rl_colors.HexColor("#1e1e2a"); BORDER=rl_colors.HexColor("#2a2a36")
    def st(name,**kw): return ParagraphStyle(name,**kw)
    S_TITLE=st("t",fontSize=18,textColor=ACCENT,fontName="Helvetica-Bold")
    S_SEC=st("s",fontSize=11,textColor=ACCENT,fontName="Helvetica-Bold",spaceAfter=2)
    S_LABEL=st("l",fontSize=9,textColor=MUTED,fontName="Helvetica")
    S_VAL=st("v",fontSize=10,textColor=TEXT,fontName="Helvetica")
    S_NOTE=st("n",fontSize=8,textColor=MUTED,fontName="Helvetica-Oblique",leftIndent=10)
    job=data.get("job",{}); history=data.get("history",{}); colors_d=data.get("colors",{})
    supplier=data.get("supplier",{}); client=data.get("client",{}); result_d=data.get("result",{})
    W=18*cm; story=[]
    def yn(v): return "Evet" if v is True else "Hayir" if v is False else "-"
    def sec(title):
        story.append(Spacer(1,0.25*cm))
        t=Table([[Paragraph(title,S_SEC)]],colWidths=[W])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),SEC_BG),("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("LINEBELOW",(0,0),(-1,-1),1,ACCENT)]))
        story.append(t)
    def row(label,value,status=None):
        sc_p=""
        if status is True:  sc_p=Paragraph("OK",st("ok",fontSize=9,fontName="Helvetica-Bold",textColor=OK_C))
        if status is False: sc_p=Paragraph("NOK",st("nk",fontSize=9,fontName="Helvetica-Bold",textColor=NOK_C))
        t=Table([[Paragraph(label,S_LABEL),Paragraph(str(value or "-"),S_VAL),sc_p]],colWidths=[5.5*cm,10.5*cm,2*cm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),CARD),("LINEBELOW",(0,0),(-1,-1),0.5,BORDER),("LEFTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
        story.append(t)
    t=Table([[Paragraph("SGK MAKEREADY CHECKLIST",S_TITLE),Paragraph(data.get("date",""),st("d",fontSize=10,textColor=MUTED,fontName="Helvetica",alignment=TA_RIGHT))]],colWidths=[13*cm,5*cm])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story.append(t); story.append(HRFlowable(width=W,thickness=2,color=ACCENT,spaceAfter=4))
    sec("1. IS BILGILERI")
    for lbl,key in [("Marka","brand"),("Variant","variant"),("Job No","jobNo"),("Tedarikci","supplier"),("Baski Teknigi","technique"),("Substrate","substrate"),("SGK Attendee","sgkAttendee")]:
        row(lbl,job.get(key,""))
    sec("2. IS GECMISI")
    hp=history.get("hasPrevious"); row("Daha once basildi mi?",yn(hp),hp)
    if hp is True:
        rsa=history.get("refSampleAvailable"); row("Referans numune mevcut?",yn(rsa),rsa)
        if rsa is True:
            row("Numune No",history.get("refSampleNo","")); row("Numune Tarihi",history.get("refSampleDate",""))
        mc=history.get("majorChange")
        row("Major degisiklik?",yn(mc),False if mc is True else True if mc is False else None)
        if mc is True: row("Degisiklik aciklamasi",history.get("majorChangeDesc",""))
        issues=history.get("prevIssues",[])
        if issues:
            sec("   Onceki Baski Sorunlari")
            for i,iss in enumerate(issues):
                n=iss.get("notified")
                row(str(i+1)+". Sorun",iss.get("note",""),True if n is True else False if n is False else None)
    sec("3. RENK REFERANSLARI")
    gp=colors_d.get("gmgProfile"); gp_name=colors_d.get("gmgProfileName","")
    row("GMG / ICC Profili",("Tanimli"+((" - "+gp_name) if gp_name else "")) if gp is True else "Tanimsiz" if gp is False else "-",gp)
    spots=colors_d.get("spotColors",[])
    if spots:
        hdr=[Paragraph(h,st("sh",fontSize=9,fontName="Helvetica-Bold",textColor=ACCENT)) for h in ["Renk Adi","Pantone","L* a* b*","dE"]]
        rd=[hdr]+[[Paragraph(sc.get(k,"-"),S_VAL) for k in ["name","pantone","lab","tolerance"]] for sc in spots]
        t=Table(rd,colWidths=[4.5*cm,4.5*cm,5.5*cm,3.5*cm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),SEC_BG),("BACKGROUND",(0,1),(-1,-1),CARD),("LINEBELOW",(0,0),(-1,-1),0.5,BORDER),("LEFTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        story.append(t)
    sec("4. TEDARIKCI HAZIRLIK")
    sv=supplier.get("substrateConfirmed"); row("Substrate teyit edildi",yn(sv),sv)
    if supplier.get("substrateNote"): story.append(Paragraph("not: "+supplier["substrateNote"],S_NOTE))
    tm=supplier.get("techniqueMatch"); row("Baski teknigi repro ile uyusuyor",yn(tm),tm)
    if supplier.get("techniqueNote"): story.append(Paragraph("not: "+supplier["techniqueNote"],S_NOTE))
    ps=supplier.get("proofSent"); row("Proof / referans baski iletildi",yn(ps),ps)
    if ps is True: row("Proof Durumu",supplier.get("proofStatus",""))
    if supplier.get("proofSentNote"): story.append(Paragraph("not: "+supplier["proofSentNote"],S_NOTE))
    sec("5. MUSTERI ONAYLARI")
    pa=client.get("proofApproved"); row("Musteri proof onayi alindi",yn(pa),pa)
    if client.get("proofApprovedNote"): story.append(Paragraph("not: "+client["proofApprovedNote"],S_NOTE))
    pin=client.get("prevIssuesNotified"); row("Onceki sorunlar musteriye iletildi",yn(pin),pin)
    if client.get("prevIssuesNotifiedNote"): story.append(Paragraph("not: "+client["prevIssuesNotifiedNote"],S_NOTE))
    sec("6. SONUC")
    if result_d.get("notes"): row("Notlar",result_d["notes"])
    decision=result_d.get("decision")
    dec_text="BASKIYA HAZIR" if decision=="ready" else "HAZIR DEGIL" if decision=="notready" else "KARAR VERILMEDI"
    dec_fg=OK_C if decision=="ready" else NOK_C if decision=="notready" else MUTED
    dec_bg=OK_BG if decision=="ready" else NOK_BG if decision=="notready" else CARD
    story.append(Spacer(1,0.3*cm))
    t=Table([[Paragraph(dec_text,st("dc",fontSize=16,fontName="Helvetica-Bold",textColor=dec_fg,alignment=TA_CENTER))]],colWidths=[W])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),dec_bg),("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12)]))
    story.append(t)
    doc.build(story); return buf.getvalue()


def generate_press_pdf(data):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=1.5*cm)
    DARK=rl_colors.HexColor("#0f0f12"); CARD=rl_colors.HexColor("#1c1c24"); ACCENT=rl_colors.HexColor("#e8c84a")
    TEXT=rl_colors.HexColor("#e8e8f0"); MUTED=rl_colors.HexColor("#9898b0"); OK_C=rl_colors.HexColor("#3ecf8e")
    SEC_BG=rl_colors.HexColor("#1e1e2a"); BORDER=rl_colors.HexColor("#2a2a36")
    PULL_COLORS=["#1e2a3a","#1e2e2a","#2a1e2e","#2a2a1e","#1e2a2a","#2e1e2a"]
    def st(name,**kw): return ParagraphStyle(name,**kw)
    S_TITLE=st("t",fontSize=18,textColor=ACCENT,fontName="Helvetica-Bold")
    S_SEC=st("s",fontSize=11,textColor=ACCENT,fontName="Helvetica-Bold",spaceAfter=2)
    S_LABEL=st("l",fontSize=9,textColor=MUTED,fontName="Helvetica")
    S_VAL=st("v",fontSize=10,textColor=TEXT,fontName="Helvetica")
    W=18*cm; story=[]
    def sec(title):
        story.append(Spacer(1,0.25*cm))
        t=Table([[Paragraph(title,S_SEC)]],colWidths=[W])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),SEC_BG),("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("LINEBELOW",(0,0),(-1,-1),1,ACCENT)]))
        story.append(t)
    def row(label,value):
        t=Table([[Paragraph(label,S_LABEL),Paragraph(str(value or "-"),S_VAL)]],colWidths=[5.5*cm,12.5*cm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),CARD),("LINEBELOW",(0,0),(-1,-1),0.5,BORDER),("LEFTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
        story.append(t)
    t=Table([[Paragraph("SGK PRESS APPROVAL REPORT",S_TITLE),Paragraph(data.get("date",""),st("d",fontSize=10,textColor=MUTED,fontName="Helvetica",alignment=TA_RIGHT))]],colWidths=[13*cm,5*cm])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story.append(t); story.append(HRFlowable(width=W,thickness=2,color=ACCENT,spaceAfter=4))
    sec("GENEL BILGILER")
    for lbl,key in [("Tarih","date"),("Saat","time"),("Printer","printer"),("Adres","printerAddress"),("Printer Attendee","printerAttendee"),("SGK Attendee","sgkAttendee"),("Client Attendee","clientAttendee"),("Marka","brand"),("Variant","variant"),("Client Product Ref.","clientProductRef"),("SGK Job Number","sgkJobNumber")]:
        row(lbl,data.get(key,""))
    if data.get("notes"): row("Notlar",data["notes"])
    if data.get("jobPreview"): row("Job Preview",data["jobPreview"])
    sec("ON-PRESS ADJUSTMENT")
    pulls=data.get("pulls",[])
    for i,pull in enumerate(pulls):
        bg=rl_colors.HexColor(PULL_COLORS[i%len(PULL_COLORS)])
        story.append(Spacer(1,0.15*cm))
        t=Table([[Paragraph("Pull #"+str(i+1),st("p",fontSize=10,fontName="Helvetica-Bold",textColor=ACCENT))]],colWidths=[W])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),bg),("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        story.append(t)
        row("Saat",pull.get("time",""))
        row("Yorumlar",pull.get("comments",""))
        for j,adj in enumerate(pull.get("adjustments",[])):
            if adj: row("Ayarlama "+str(j+1),adj)
    ap=data.get("approvedPull")
    sec("ONAY")
    row("Onaylanan Pull",str(ap+1) if ap is not None else "-")
    row("Onay Saati",data.get("approvalTime",""))
    row("Onay Yorumu",data.get("approvalComments",""))
    if data.get("printSampleComments") or data.get("issuesAnalysis") or data.get("improvements"):
        sec("PRINT SAMPLE ANALYSIS")
        if data.get("printSampleComments"): row("Yorumlar",data["printSampleComments"])
        if data.get("issuesAnalysis"): row("Issues Analysis",data["issuesAnalysis"])
        if data.get("improvements"): row("Improvements",data["improvements"])
    doc.build(story); return buf.getvalue()


@app.route("/generate-pdf-makeready", methods=["POST","OPTIONS"])
def generate_pdf_makeready():
    if request.method=="OPTIONS": return cors(Response())
    data=request.get_json()
    pdf=generate_makeready_pdf(data)
    job=data.get("job",{}); brand=job.get("brand","Report").replace(" ","_")
    date=data.get("date","").replace(".","-")
    return cors(send_file(io.BytesIO(pdf),as_attachment=True,
        download_name="Makeready_"+brand+"_"+date+".pdf",mimetype="application/pdf"))

@app.route("/generate-pdf-press", methods=["POST","OPTIONS"])
def generate_pdf_press():
    if request.method=="OPTIONS": return cors(Response())
    data=request.get_json()
    pdf=generate_press_pdf(data)
    brand=data.get("brand","Report").replace(" ","_")
    date=data.get("date","").replace(".","-")
    return cors(send_file(io.BytesIO(pdf),as_attachment=True,
        download_name="Press_Approval_"+brand+"_"+date+".pdf",mimetype="application/pdf"))


if __name__=="__main__":
    port=int(os.environ.get("PORT",5050))
    app.run(host="0.0.0.0",port=port)
